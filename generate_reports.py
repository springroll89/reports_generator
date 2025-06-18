import os
import sys
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from matplotlib.font_manager import FontProperties
import traceback
from datetime import datetime
import xlsxwriter
from openpyxl import load_workbook

# 确保使用相对路径，适用于Windows环境
def get_base_dir():
    """获取基础目录，兼容PyInstaller打包后的路径"""
    if getattr(sys, 'frozen', False):
        # 打包后的程序运行路径
        return os.path.dirname(sys.executable)
    else:
        # 开发环境下运行路径
        return os.path.dirname(os.path.abspath(__file__))

# 设置基础路径
BASE_DIR = get_base_dir()
INPUT_DIR = os.path.join(BASE_DIR, 'reports_input')
OUTPUT_DIR = os.path.join(BASE_DIR, 'reports_output')

# 确保输出目录存在
if not os.path.exists(OUTPUT_DIR):
    os.makedirs(OUTPUT_DIR)

def setup_chinese_font():
    """设置中文字体，多种备选方案"""
    try:
        # 尝试使用自带的仿宋字体
        font_file = os.path.join(INPUT_DIR, '.仿宋_GB2312_0.TTF')
        if os.path.exists(font_file):
            prop = FontProperties(fname=font_file)
            plt.rcParams['font.family'] = prop.get_name()
            return prop
        else:
            # Windows系统字体备选
            plt.rcParams['font.sans-serif'] = ['SimSun', 'Microsoft YaHei', 'SimHei', 'Arial Unicode MS']
            plt.rcParams['axes.unicode_minus'] = False
            return FontProperties(family='SimSun')
    except Exception as e:
        print(f"字体加载失败: {str(e)}")
        # 使用默认字体
        return None

def load_reference_curve():
    """加载基准曲线数据"""
    try:
        curve_file = os.path.join(INPUT_DIR, '.基准曲线 1.csv')
        if not os.path.exists(curve_file):
            print(f"错误: 找不到基准曲线文件: {curve_file}")
            return None
        
        # 尝试不同的编码方式读取CSV
        encodings = ['utf-8', 'gbk', 'gb2312', 'latin1']
        df = None
        
        for encoding in encodings:
            try:
                df = pd.read_csv(curve_file, encoding=encoding)
                break
            except UnicodeDecodeError:
                continue
        
        if df is None:
            print("错误: 无法读取基准曲线文件，所有编码方式均失败")
            return None
            
        return df
    except Exception as e:
        print(f"加载基准曲线数据失败: {str(e)}")
        return None

def load_template_excel():
    """加载Excel模板文件"""
    try:
        template_file = os.path.join(INPUT_DIR, '.副本4P、2P22M氧烛分层及刻度.xlsx')
        if not os.path.exists(template_file):
            print(f"错误: 找不到Excel模板文件: {template_file}")
            return None
            
        try:
            wb = load_workbook(template_file)
            return wb
        except Exception as e:
            print(f"加载Excel模板失败: {str(e)}")
            return None
    except Exception as e:
        print(f"加载Excel模板文件失败: {str(e)}")
        return None

def process_data(data):
    """处理数据示例"""
    # 这里是根据你的具体需求处理数据
    # 示例：计算一些简单的统计数据
    processed_data = {
        'mean': np.mean(data),
        'std': np.std(data),
        'min': np.min(data),
        'max': np.max(data)
    }
    return processed_data

def generate_plot(data, title, output_path):
    """生成图表"""
    try:
        font_prop = setup_chinese_font()
        
        plt.figure(figsize=(10, 6))
        plt.plot(data, marker='o')
        
        if font_prop:
            plt.title(title, fontproperties=font_prop)
            plt.xlabel('测量点', fontproperties=font_prop)
            plt.ylabel('值', fontproperties=font_prop)
        else:
            plt.title(title)
            plt.xlabel('测量点')
            plt.ylabel('值')
            
        plt.grid(True)
        plt.tight_layout()
        
        # 确保保存路径存在
        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        plt.savefig(output_path)
        plt.close()
        return True
    except Exception as e:
        print(f"生成图表失败: {str(e)}")
        return False

def generate_excel_report(data, output_path):
    """生成Excel报告"""
    try:
        # 创建一个新的Excel工作簿和工作表
        workbook = xlsxwriter.Workbook(output_path)
        worksheet = workbook.add_worksheet('报告数据')
        
        # 添加标题格式
        title_format = workbook.add_format({
            'bold': True,
            'font_size': 14,
            'align': 'center',
            'valign': 'vcenter'
        })
        
        header_format = workbook.add_format({
            'bold': True,
            'bg_color': '#D9E1F2',
            'border': 1
        })
        
        # 添加单元格格式
        cell_format = workbook.add_format({
            'border': 1
        })
        
        # 写入标题
        worksheet.merge_range('A1:D1', '数据分析报告', title_format)
        
        # 写入日期
        worksheet.write(1, 0, '生成日期:', header_format)
        worksheet.write(1, 1, datetime.now().strftime('%Y-%m-%d %H:%M:%S'), cell_format)
        
        # 写入表头
        headers = ['指标', '值']
        for col, header in enumerate(headers):
            worksheet.write(3, col, header, header_format)
        
        # 写入数据
        row = 4
        for key, value in data.items():
            worksheet.write(row, 0, key, cell_format)
            worksheet.write(row, 1, value, cell_format)
            row += 1
        
        # 设置列宽
        worksheet.set_column('A:A', 15)
        worksheet.set_column('B:B', 20)
        
        # 关闭工作簿
        workbook.close()
        return True
    except Exception as e:
        print(f"生成Excel报告失败: {str(e)}")
        return False

def main():
    """主函数"""
    try:
        print("报告生成器启动...")
        print(f"基础目录: {BASE_DIR}")
        print(f"输入目录: {INPUT_DIR}")
        print(f"输出目录: {OUTPUT_DIR}")
        
        # 加载基准曲线数据
        curve_data = load_reference_curve()
        if curve_data is None:
            raise Exception("无法加载基准曲线数据，程序终止")
        
        print("成功加载基准曲线数据")
        
        # 处理简单的示例数据 - 在实际应用中修改为你的数据处理逻辑
        sample_data = curve_data.iloc[:, 1].values if len(curve_data.columns) > 1 else np.random.rand(10)
        processed_data = process_data(sample_data)
        
        # 生成报表的时间戳
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        
        # 生成图表
        plot_path = os.path.join(OUTPUT_DIR, f'报告图表_{timestamp}.png')
        if not generate_plot(sample_data, '数据趋势分析', plot_path):
            print("警告: 图表生成失败")
        else:
            print(f"成功生成图表: {plot_path}")
        
        # 生成Excel报告
        excel_path = os.path.join(OUTPUT_DIR, f'数据报告_{timestamp}.xlsx')
        if not generate_excel_report(processed_data, excel_path):
            print("警告: Excel报告生成失败")
        else:
            print(f"成功生成Excel报告: {excel_path}")
        
        print("报告生成完成!")
        return True
    except Exception as e:
        print(f"错误: {str(e)}")
        print("详细错误信息:")
        traceback.print_exc()
        
        # 将错误信息写入日志文件
        try:
            error_log = os.path.join(OUTPUT_DIR, 'error_log.txt')
            with open(error_log, 'w', encoding='utf-8') as f:
                f.write(f"错误发生时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
                f.write(f"错误信息: {str(e)}\n")
                f.write("详细错误堆栈:\n")
                f.write(traceback.format_exc())
            print(f"错误日志已保存到: {error_log}")
        except:
            print("保存错误日志失败")
        
        return False

# 如果作为脚本运行
if __name__ == "__main__":
    success = main()
    
    # 如果是直接运行的EXE，在结束时暂停
    if getattr(sys, 'frozen', False):
        input("按Enter键退出...")
    
    sys.exit(0 if success else 1)
